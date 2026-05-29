import csv
import json
import os
import unicodedata
from datetime import datetime, timedelta

import openpyxl

ARCHIVO = os.getenv("ARCHIVO_HISTORICOS", "trazabilidad.tsv")
EXCEL_SENADORES = os.getenv("EXCEL_SENADORES", "Senadores 2026.xlsx")

TIPOS = {
    "PL": "Proyecto de Ley",
    "PD": "Proyecto de Declaración",
    "PC": "Proyecto de Comunicación",
    "PR": "Proyecto de Resolución",
    "CA": "Com. de Auditoría",
    "AC": "Acuerdo",
    "CV": "Com. Varias",
}


def normalizar(s):
    """Convierte a mayúsculas y elimina acentos/diacríticos."""
    s = str(s).upper().strip()
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def cargar_senadores():
    """Lee Senadores 2026.xlsx y devuelve {apellido_normalizado: {bloque, provincia}}."""
    dicc = {}
    try:
        wb = openpyxl.load_workbook(EXCEL_SENADORES, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = dict(zip(headers, row))
            apellido = r.get("APELLIDO") or ""
            clave = normalizar(apellido)
            if clave:
                dicc[clave] = {
                    "bloque": str(r.get("BLOQUE") or "").strip(),
                    "provincia": str(r.get("PROVINCIA") or "").strip(),
                }
        wb.close()
    except FileNotFoundError:
        print(f"Advertencia: no se encontró {EXCEL_SENADORES}, bloque/provincia quedarán vacíos.")
    return dicc


def parse_fecha_mesa(valor):
    """Extrae DD/MM/YYYY del campo MESA que tiene formato 'DD/MM/YYYY -'."""
    parte = valor.strip().split(" ")[0]
    try:
        return datetime.strptime(parte, "%d/%m/%Y")
    except ValueError:
        return None


def apellido_principal(autor):
    """Extrae el apellido del primer firmante: todo lo que está antes de la primera coma."""
    return autor.split(",")[0].strip()


def main():
    senadores = cargar_senadores()
    print(f"Senadores cargados desde Excel: {len(senadores)}")

    ahora = datetime.now()
    hace_24h = ahora - timedelta(hours=24)

    proyectos = []
    total_leidas = 0

    with open(ARCHIVO, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for fila in reader:
            total_leidas += 1
            fecha = parse_fecha_mesa(fila.get("MESA", ""))
            if fecha is None:
                continue
            if not (hace_24h <= fecha <= ahora):
                continue

            nro = fila.get("NRO", "").strip()
            anio = fila.get("ANIO", "").strip()
            origen = fila.get("ORIGEN", "").strip()
            tipo = fila.get("TIPO", "").strip()
            caratula = fila.get("CARATULA", "").strip()
            autor = fila.get("AUTOR", "").strip()

            titulo = caratula.split(":", 1)[1].strip() if ":" in caratula else caratula

            clave = normalizar(apellido_principal(autor))
            datos_senador = senadores.get(clave, {})

            comisiones = [
                fila.get(c, "").strip()
                for c in ("COM1", "COM2", "COM3", "COM4", "COM5")
                if fila.get(c, "").strip()
            ]

            anio_corto = str(anio)[-2:] if len(str(anio)) >= 2 else anio
            url = (
                f"https://www.senado.gob.ar/parlamentario/comisiones/verExp/"
                f"{nro}.{anio_corto}/{origen}/{tipo}"
            )

            proyectos.append({
                "expediente": f"{nro}/{anio}",
                "tipo": tipo,
                "tipo_label": TIPOS.get(tipo, tipo),
                "titulo": titulo,
                "autor": autor,
                "bloque": datos_senador.get("bloque", ""),
                "provincia": datos_senador.get("provincia", ""),
                "fecha": fecha.strftime("%d/%m/%Y"),
                "comisiones": " | ".join(comisiones),
                "url": url,
            })

    resultado = {
        "fecha_generacion": ahora.strftime("%d/%m/%Y %H:%M"),
        "total": len(proyectos),
        "proyectos": proyectos,
    }

    with open("resumen_diario.json", "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"Filas leídas: {total_leidas} | Filas en últimas 24hs: {len(proyectos)}")


if __name__ == "__main__":
    main()
