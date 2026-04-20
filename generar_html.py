# =============================================================================
#  GENERADOR DE REPORTE — Prosecretaría Parlamentaria
#  Senado de la Nación Argentina
# =============================================================================

# --- CONFIGURACIÓN MANUAL (solo para uso sin scraper) ------------------------

EXCEL_PROYECTOS  = "Ingresados_primera_quincena_marzo.xlsx"
EXCEL_SENADORES  = "Senadores_2026.xlsx"
TITULO_PERIODO   = "1.ª Quincena · Marzo 2026"
FECHA_DATOS      = "13/03/2026"
ARCHIVO_SALIDA   = "index.html"

# -----------------------------------------------------------------------------

import json, os, sys

TIPOS = {
    "PL": "Proyecto de Ley",
    "PD": "Proyecto de Declaración",
    "PC": "Proyecto de Comunicación",
    "PR": "Proyecto de Resolución",
    "CA": "Com. de Auditoría",
    "AC": "Acuerdo",
    "CV": "Com. Varias",
}

# ── Estilos ───────────────────────────────────────────────────────────────────

CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Poppins',Calibri,sans-serif;background:#E8EEF5;color:#4A4A4A;font-size:15px;line-height:1.5}

.header{background:#1B5EA2;padding:14px 16px;position:sticky;top:0;z-index:100;border-bottom:2px solid #0d3f73}
.header-row{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px}
.header-inst{font-size:9px;font-weight:400;color:rgba(255,255,255,0.75);text-transform:uppercase;letter-spacing:2px}
.header-dep{font-size:9px;font-weight:700;color:rgba(255,255,255,0.75);text-transform:uppercase;letter-spacing:2px}
.header-title{font-size:18px;font-weight:700;color:#fff}
.header-subtitle{font-size:12px;color:rgba(255,255,255,0.8);margin-top:1px}

.tab-bar{display:flex;background:#0d3f73;padding:0 12px;gap:2px;position:sticky;top:68px;z-index:99}
.tab-btn{padding:11px 24px;background:transparent;border:none;color:rgba(255,255,255,0.55);font-family:inherit;font-size:13px;font-weight:600;cursor:pointer;border-bottom:3px solid transparent;transition:all .2s;text-transform:uppercase;letter-spacing:1px}
.tab-btn.active{color:#fff;border-bottom-color:#fff}
.tab-btn:hover{color:rgba(255,255,255,0.85)}
.tab-content{display:none}
.tab-content.active{display:block}

/* ── Dashboard ──────────────────────────────────────────────────────── */
.section-block{background:#fff;margin:12px;border-radius:10px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.08)}
.section-header{background:#1B5EA2;padding:10px 16px;display:flex;justify-content:space-between;align-items:center}
.section-header h2{font-size:11px;font-weight:700;color:#fff;text-transform:uppercase;letter-spacing:1.5px}
.section-hint{font-size:10px;color:rgba(255,255,255,0.65)}
.section-body{padding:16px}

/* Dashboard PC: fila de stats arriba + 3 columnas iguales abajo */
.dash-stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}
.dash-panels-row{display:grid;grid-template-columns:repeat(3,1fr);gap:16px}
@media(max-width:900px){
  .dash-stats-row{grid-template-columns:repeat(2,1fr)}
  .dash-panels-row{grid-template-columns:1fr}
}
.stat-card{background:#F5F7FA;border-radius:8px;padding:14px 12px;border-left:4px solid #1B5EA2}
.stat-num{font-size:28px;font-weight:700;color:#1B5EA2;line-height:1}
.stat-label{font-size:11px;color:#4A4A4A;margin-top:3px}
.dash-subtitle{font-size:11px;font-weight:600;color:#2E75B6;text-transform:uppercase;letter-spacing:1px;margin:0 0 8px;padding-bottom:4px;border-bottom:1px solid #D6E4F0}
.tipo-bar-row{display:flex;align-items:center;gap:8px;margin-bottom:7px;cursor:pointer;padding:3px 6px;border-radius:6px;transition:background .12s}
.tipo-bar-row:hover{background:#F0F4FA}
.tipo-bar-row.on{background:#D6E4F0}
.tipo-pill{font-size:11px;font-weight:700;border-radius:4px;padding:3px 8px;min-width:36px;text-align:center;flex-shrink:0}
.tipo-nombre{font-size:12px;color:#4A4A4A;flex:1}
.bar-track{flex:2;height:7px;background:#D6E4F0;border-radius:4px;overflow:hidden}
.bar-fill{height:100%;border-radius:4px;transition:width .3s}
.tipo-count{font-size:12px;font-weight:700;min-width:28px;text-align:right}
.bloque-row{display:flex;align-items:center;gap:8px;margin-bottom:6px;cursor:pointer;padding:3px 6px;border-radius:6px;transition:background .12s}
.bloque-row:hover,.com-row:hover{background:#F0F4FA}
.bloque-row.on,.com-row.on{background:#D6E4F0}
.bloque-name{font-size:11px;color:#4A4A4A;flex:1;line-height:1.3}
.bloque-bar-track{flex:2;height:6px;background:#D6E4F0;border-radius:3px;overflow:hidden}
.bloque-bar-fill{height:100%;border-radius:3px;transition:width .3s}
.bloque-count{font-size:11px;font-weight:700;color:#2E75B6;min-width:24px;text-align:right}
.com-row{display:flex;align-items:center;gap:8px;margin-bottom:6px;cursor:pointer;padding:3px 6px;border-radius:6px;transition:background .12s}
.com-name{font-size:11px;color:#4A4A4A;flex:1;line-height:1.3}
.com-bar-track{flex:2;height:6px;background:#D6E4F0;border-radius:3px;overflow:hidden}
.com-bar-fill{height:100%;background:#2E75B6;border-radius:3px;transition:width .3s}
.com-count{font-size:11px;font-weight:700;color:#2E75B6;min-width:24px;text-align:right}
.dash-context{font-size:11px;color:#2E75B6;background:#EAF0FA;border-radius:6px;padding:6px 10px;margin-bottom:10px;display:none}
.dash-context.visible{display:block}

/* ── Detalle: layout dos columnas ───────────────────────────────────── */
.detalle-layout{display:flex;gap:16px;padding:12px;align-items:flex-start}
.filters-panel{width:280px;flex-shrink:0;background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,0.08);overflow-y:auto;max-height:calc(100vh - 140px);position:sticky;top:120px}
.filters-panel .section-header{border-radius:0}
.filters-body{padding:14px}
.results-panel{flex:1;min-width:0}
@media(max-width:900px){
  .detalle-layout{flex-direction:column}
  .filters-panel{width:100%;position:static}
}

.search-box{width:100%;padding:10px 12px;border:1.5px solid #D6E4F0;border-radius:8px;font-family:inherit;font-size:13px;color:#4A4A4A;outline:none;margin-bottom:10px;background:#fff}
.search-box:focus{border-color:#1B5EA2}
.filter-label{font-size:10px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:1px;margin-bottom:5px;margin-top:10px}
.filter-label:first-child{margin-top:0}
.filter-row{display:flex;gap:5px;flex-wrap:wrap;margin-bottom:4px}
.chip{padding:6px 11px;border-radius:20px;border:1.5px solid #D6E4F0;background:#fff;font-family:inherit;font-size:11px;color:#4A4A4A;cursor:pointer;transition:all .15s;white-space:nowrap;-webkit-appearance:none;line-height:1.2}
.chip.on{background:#1B5EA2;border-color:#1B5EA2;color:#fff;font-weight:600}
.results-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;flex-wrap:wrap;gap:8px}
.results-count{font-size:12px;color:#888}
.btn-export{padding:7px 14px;border-radius:8px;border:1.5px solid #1B5EA2;background:#fff;color:#1B5EA2;font-family:inherit;font-size:11px;font-weight:600;cursor:pointer;transition:all .15s}
.btn-export:hover{background:#1B5EA2;color:#fff}
.select-wrapper{position:relative;display:block;margin-bottom:4px}
.filter-select{width:100%;padding:8px 32px 8px 11px;border:1.5px solid #D6E4F0;border-radius:8px;font-family:inherit;font-size:12px;color:#4A4A4A;background:#fff;outline:none;cursor:pointer;-webkit-appearance:none;appearance:none;transition:border-color .15s}
.filter-select:focus,.filter-select.on{border-color:#1B5EA2;background:#EAF0FA;color:#1B5EA2;font-weight:600}
.select-arrow{position:absolute;right:10px;top:50%;transform:translateY(-50%);pointer-events:none;color:#888;font-size:12px}
.date-range{display:flex;flex-direction:column;gap:5px;margin-bottom:4px}
.date-input{width:100%;padding:7px 10px;border:1.5px solid #D6E4F0;border-radius:8px;font-family:inherit;font-size:12px;color:#4A4A4A;background:#fff;outline:none}
.date-input:focus{border-color:#1B5EA2}
.date-sep{font-size:10px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:1px}

.card{background:#fff;border-radius:10px;margin-bottom:10px;overflow:hidden;border:1px solid #D6E4F0;box-shadow:0 1px 3px rgba(0,0,0,0.05)}
.card-exp{display:flex;align-items:center;justify-content:space-between;padding:9px 14px 7px;border-bottom:1px solid #EEF2F8;background:#F5F8FC}
.exp-id{display:flex;align-items:center;gap:8px}
.exp-badge{font-size:11px;font-weight:700;padding:4px 9px;border-radius:4px;flex-shrink:0;letter-spacing:.5px}
.exp-nro{font-size:14px;font-weight:700;color:#1B5EA2}
.exp-link{font-size:11px;color:#2E75B6;text-decoration:none;font-weight:600;border:1px solid #2E75B6;padding:3px 9px;border-radius:12px;white-space:nowrap;transition:all .15s}
.exp-link:hover{background:#2E75B6;color:#fff}
.exp-fecha{font-size:11px;color:#888}
.card-body{padding:12px 14px 6px}
.extracto{font-size:14px;font-weight:600;color:#2C2C2C;line-height:1.4;margin-bottom:10px}
.card-meta{display:flex;flex-direction:column;gap:5px;padding-bottom:10px}
.meta-row{display:flex;gap:6px;align-items:flex-start;flex-wrap:wrap}
.meta-bold{font-size:13px;font-weight:600;color:#4A4A4A}
.btag{display:inline-block;font-size:11px;font-weight:600;padding:3px 8px;border-radius:4px;margin-right:4px;margin-bottom:3px}
.ctag{display:inline-block;font-size:11px;padding:3px 8px;border-radius:4px;margin-right:4px;margin-bottom:3px;background:#EAF0FA;color:#1B5EA2;border:1px solid #c8daf0}
.no-results{text-align:center;padding:48px 16px;color:#aaa;font-size:14px}
.footer{text-align:center;padding:20px 16px;font-size:11px;color:#aaa;font-style:italic}

/* SEMÁNTICO: desactivado temporalmente - reactivar cuando se retome
.btn-sem{padding:10px 12px;border-radius:8px;border:1.5px solid #1B5EA2;background:#1B5EA2;color:#fff;font-family:inherit;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap;transition:all .15s;flex-shrink:0}
.btn-sem:hover:not(:disabled){background:#0d3f73;border-color:#0d3f73}
.btn-sem:disabled{opacity:0.5;cursor:not-allowed}
.sem-status{font-size:11px;color:#888;margin-top:3px;min-height:16px}
.sem-badge{font-size:10px;font-weight:700;padding:2px 7px;border-radius:10px;background:#E8F4E8;color:#1a7a4a;border:1px solid #b8e0b8;flex-shrink:0;align-self:center}
.sem-banner{background:#EAF9EA;border:1px solid #b8e0b8;border-radius:8px;padding:8px 12px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:center;font-size:12px;color:#1a7a4a;flex-wrap:wrap;gap:6px}
.sem-divider{border:none;border-top:1.5px solid #D6E4F0;margin:10px 0}
*/
"""

JS = r"""
var TIPOS={PL:'Proy. de Ley',PD:'Declaraci\u00f3n',PC:'Comunicaci\u00f3n',PR:'Resoluci\u00f3n',CA:'Com. Auditor\u00eda',AC:'Acuerdo',CV:'Com. Varias'};
var TIPO_FG={PL:'#1B5EA2',PD:'#2E75B6',PC:'#0d7a4a',PR:'#5B4DA0',CA:'#1a7a4a',AC:'#7a5c1a',CV:'#7a1a3a'};
var TIPO_BG={PL:'#D6E4F0',PD:'#EAF0FA',PC:'#DCF0E8',PR:'#EDE8FA',CA:'#E0F4EC',AC:'#F9F0DA',CV:'#FAE0EA'};
var BC=['#1B5EA2','#2E75B6','#5B4DA0','#1a7a4a','#7a5c1a','#7a1a3a','#2E8B7A','#6B3A2A','#1a4a7a','#4a7a1a','#7a1a5a','#2a7a6a','#5a2a7a','#2a5a2a'];
var ALL_BLOQUES=[];
var dashFiltroTipo='',dashFiltroBloque='',dashFiltroCom='',dashActiveAnio='';
var activeTipos={},activeBloque='',activeOrigen='',activeProvincia='',activeAnio='';
var semActivo=false,semResultados=[],EMBEDDINGS=null,semPipeline=null;

var histInited=false;
function switchTab(id){
  document.querySelectorAll('.tab-btn').forEach(function(b){b.classList.remove('active')});
  document.querySelectorAll('.tab-content').forEach(function(c){c.classList.remove('active')});
  document.getElementById('tab-'+id).classList.add('active');
  document.querySelector('[data-tab="'+id+'"]').classList.add('active');
  if(id==='historico'&&!histInited){histInited=true;initHist();}
}

function init(){
  var bset={};
  DATA.forEach(function(p){p.bloques.forEach(function(b){if(b)bset[b]=1})});
  ALL_BLOQUES=Object.keys(bset).sort();

  var cset1={},csetAdic={};
  DATA.forEach(function(p){
    if(p.comisiones[0])cset1[p.comisiones[0]]=1;
    if(p.comisiones[1])csetAdic[p.comisiones[1]]=1;
    if(p.comisiones[2])csetAdic[p.comisiones[2]]=1;
  });
  var cSel1=document.getElementById('com-select-1');
  Object.keys(cset1).sort().forEach(function(c){
    var o=document.createElement('option');o.value=c;o.textContent=c;cSel1.appendChild(o);
  });
  var cSelAdic=document.getElementById('com-select-adic');
  Object.keys(csetAdic).sort().forEach(function(c){
    var o=document.createElement('option');o.value=c;o.textContent=c;cSelAdic.appendChild(o);
  });

  var aset={};
  DATA.forEach(function(p){p.autores.forEach(function(a){aset[a]=1})});
  var aSel=document.getElementById('autor-select');
  Object.keys(aset).sort().forEach(function(a){
    var o=document.createElement('option');o.value=a;o.textContent=a;aSel.appendChild(o);
  });

  var bSel=document.getElementById('bloque-select');
  ALL_BLOQUES.forEach(function(b){
    var o=document.createElement('option');o.value=b;o.textContent=b;bSel.appendChild(o);
  });

  var provSet={};
  DATA.forEach(function(p){(p.provincias||[]).forEach(function(pv){if(pv)provSet[pv]=1})});
  var provSel=document.getElementById('provincia-select');
  Object.keys(provSet).sort().forEach(function(pv){
    var o=document.createElement('option');o.value=pv;o.textContent=pv;provSel.appendChild(o);
  });

  /* SEMÁNTICO: desactivado temporalmente - reactivar cuando se retome */

  renderDash(DATA);
  renderFilters();
  renderList();
}

function getBloqueColor(b){return BC[ALL_BLOQUES.indexOf(b)%BC.length]}

/* ── Dashboard ─────────────────────────────────────────────────── */
function getDashFiltered(){
  return DATA.filter(function(p){
    if(dashActiveAnio&&String(p.anio)!==dashActiveAnio)return false;
    if(dashFiltroTipo&&p.tipo!==dashFiltroTipo)return false;
    if(dashFiltroBloque&&p.bloques.indexOf(dashFiltroBloque)<0)return false;
    if(dashFiltroCom&&p.comisiones.indexOf(dashFiltroCom)<0)return false;
    return true;
  });
}
function calcStats(data){
  var t={},b={},c={};
  data.forEach(function(p){
    t[p.tipo]=(t[p.tipo]||0)+1;
    p.bloques.forEach(function(x){b[x]=(b[x]||0)+1});
    p.comisiones.forEach(function(x){c[x]=(c[x]||0)+1});
  });
  return{tipos:t,bloques:b,coms:c};
}
function renderDash(data){
  var s=calcStats(data),total=data.length;
  document.getElementById('stat-total').innerHTML=total;
  document.getElementById('stat-pl').innerHTML=s.tipos['PL']||0;
  document.getElementById('stat-pd').innerHTML=s.tipos['PD']||0;
  document.getElementById('stat-otros').innerHTML=total-(s.tipos['PL']||0)-(s.tipos['PD']||0);

  var partes=[];
  if(dashActiveAnio)partes.push('A\u00f1o: '+dashActiveAnio);
  if(dashFiltroTipo)partes.push('Tipo: '+(TIPOS[dashFiltroTipo]||dashFiltroTipo));
  if(dashFiltroBloque)partes.push('Bloque: '+dashFiltroBloque);
  if(dashFiltroCom)partes.push('Comisi\u00f3n: '+dashFiltroCom);
  var ctx=document.getElementById('dash-context');
  if(partes.length){
    ctx.innerHTML='Filtrando: <strong>'+partes.join(' &middot; ')+'</strong> &nbsp;<button onclick="clearDash()" style="background:none;border:none;color:#1B5EA2;cursor:pointer;font-size:11px;font-weight:700;padding:0 4px">&#x2715;</button>';
    ctx.className='dash-context visible';
  }else{ctx.className='dash-context'}

  var tipoOrder=['PL','PD','PC','PR','CA','AC','CV'],maxT=0;
  tipoOrder.forEach(function(t){if((s.tipos[t]||0)>maxT)maxT=s.tipos[t]||0});
  var tb='';
  tipoOrder.forEach(function(t){
    if(!DATA.some(function(p){return p.tipo===t}))return;
    var n=s.tipos[t]||0,pct=maxT?Math.round(n/maxT*100):0;
    var fg=TIPO_FG[t]||'#888',bg=TIPO_BG[t]||'#eee';
    var on=dashFiltroTipo===t?' on':'';
    tb+='<div class="tipo-bar-row'+on+'" onclick="clickDashTipo(\''+t+'\')"><span class="tipo-pill" style="background:'+bg+';color:'+fg+'">'+t+'</span><span class="tipo-nombre">'+(TIPOS[t]||t)+'</span><div class="bar-track"><div class="bar-fill" style="width:'+pct+'%;background:'+fg+'"></div></div><span class="tipo-count" style="color:'+fg+'">'+n+'</span></div>';
  });
  document.getElementById('tipo-bars').innerHTML=tb;

  var blist=Object.keys(s.bloques).sort(function(a,b){return s.bloques[b]-s.bloques[a]});
  var maxB=blist.length?s.bloques[blist[0]]:1;
  var bb='';
  blist.forEach(function(b){
    var n=s.bloques[b],pct=Math.round(n/maxB*100),color=getBloqueColor(b);
    var safe=b.replace(/\\/g,'\\\\').replace(/'/g,"\\'");
    var on=dashFiltroBloque===b?' on':'';
    bb+='<div class="bloque-row'+on+'" onclick="clickDashBloque(\''+safe+'\')"><span class="bloque-name">'+b+'</span><div class="bloque-bar-track"><div class="bloque-bar-fill" style="width:'+pct+'%;background:'+color+'"></div></div><span class="bloque-count">'+n+'</span></div>';
  });
  document.getElementById('bloque-bars').innerHTML=bb;

  var clist=Object.keys(s.coms).sort(function(a,b){return s.coms[b]-s.coms[a]}).slice(0,10);
  var maxC=clist.length?s.coms[clist[0]]:1;
  var cb='';
  clist.forEach(function(c){
    var n=s.coms[c],pct=Math.round(n/maxC*100);
    var safe=c.replace(/\\/g,'\\\\').replace(/'/g,"\\'");
    var on=dashFiltroCom===c?' on':'';
    cb+='<div class="com-row'+on+'" onclick="clickDashCom(\''+safe+'\')"><span class="com-name">'+c+'</span><div class="com-bar-track"><div class="com-bar-fill" style="width:'+pct+'%"></div></div><span class="com-count">'+n+'</span></div>';
  });
  document.getElementById('com-bars').innerHTML=cb;
}
function clickDashTipo(t){dashFiltroTipo=dashFiltroTipo===t?'':t;renderDash(getDashFiltered())}
function clickDashBloque(b){dashFiltroBloque=dashFiltroBloque===b?'':b;renderDash(getDashFiltered())}
function clickDashCom(c){dashFiltroCom=dashFiltroCom===c?'':c;renderDash(getDashFiltered())}
function setDashAnio(anio){
  dashActiveAnio=anio;
  ['all','2025','2026'].forEach(function(a){
    var el=document.getElementById('dash-anio-'+(a==='all'?'all':a));
    if(el)el.className='chip'+(anio===(a==='all'?'':a)?' on':'');
  });
  renderDash(getDashFiltered());
}
function clearDash(){dashFiltroTipo='';dashFiltroBloque='';dashFiltroCom='';renderDash(getDashFiltered())}

/* ── Filtros ───────────────────────────────────────────────────── */
function renderFilters(){
  var tset={};
  DATA.forEach(function(p){tset[p.tipo]=1});
  var anyT=Object.keys(activeTipos).length===0;
  var h='<button class="chip'+(anyT?' on':'')+'" onclick="toggleTipo(\'__all__\')">Todos</button>';
  Object.keys(tset).sort().forEach(function(t){
    h+='<button class="chip'+(activeTipos[t]?' on':'')+'" onclick="toggleTipo(\''+t+'\')">'+t+' &middot; '+(TIPOS[t]||t)+'</button>';
  });
  document.getElementById('tipo-filters').innerHTML=h;

  var ORIGEN_LABEL={S:'Senado',PE:'Poder Ejecutivo',CD:'Diputados',OV:'Otros'};
  var ohtml='<button class="chip'+(activeOrigen===''?' on':'')+'" onclick="toggleOrigen(\'\')">Todos</button>';
  var oset={};
  DATA.forEach(function(p){oset[p.origen]=1});
  Object.keys(oset).sort().forEach(function(o){
    ohtml+='<button class="chip'+(activeOrigen===o?' on':'')+'" onclick="toggleOrigen(\''+o+'\')">'+(ORIGEN_LABEL[o]||o)+'</button>';
  });
  document.getElementById('origen-filters').innerHTML=ohtml;
}
function toggleTipo(t){
  if(t==='__all__'){activeTipos={}}else{if(activeTipos[t])delete activeTipos[t];else activeTipos[t]=1}
  renderFilters();renderList();
}
function toggleOrigen(o){activeOrigen=activeOrigen===o?'':o;renderFilters();renderList()}
function setAnioDetalle(anio){
  activeAnio=anio;
  ['all','2025','2026'].forEach(function(a){
    var el=document.getElementById('anio-det-'+(a==='all'?'all':a));
    if(el)el.className='chip'+(anio===(a==='all'?'':a)?' on':'');
  });
  renderList();
}
function setBloque(val){
  activeBloque=val;
  var el=document.getElementById('bloque-select');
  if(el)el.className=val?'filter-select on':'filter-select';
  renderList();
}
function setProvincia(val){
  activeProvincia=val;
  var el=document.getElementById('provincia-select');
  if(el)el.className=val?'filter-select on':'filter-select';
  renderList();
}

/* ── Parsear fecha dd/mm/yyyy a Date ───────────────────────────── */
function parseFecha(s){
  if(!s)return null;
  var p=s.split('/');
  if(p.length!==3)return null;
  return new Date(parseInt(p[2]),parseInt(p[1])-1,parseInt(p[0]));
}

/* ── Lista ─────────────────────────────────────────────────────── */
function getFiltered(){
  var q=document.getElementById('search').value.toLowerCase().trim();
  var selCom1=document.getElementById('com-select-1').value;
  var selComAdic=document.getElementById('com-select-adic').value;
  var selAutor=document.getElementById('autor-select').value;
  var dDesde=document.getElementById('fecha-desde').value;
  var dHasta=document.getElementById('fecha-hasta').value;
  var fDesde=dDesde?new Date(dDesde):null;
  var fHasta=dHasta?new Date(dHasta+'T23:59:59'):null;

  return DATA.filter(function(p){
    if(activeAnio&&String(p.anio)!==activeAnio)return false;
    if(Object.keys(activeTipos).length&&!activeTipos[p.tipo])return false;
    if(activeBloque&&p.bloques.indexOf(activeBloque)<0)return false;
    if(activeOrigen&&p.origen!==activeOrigen)return false;
    if(activeProvincia&&(!p.provincias||p.provincias.indexOf(activeProvincia)<0))return false;
    if(selCom1&&p.comisiones[0]!==selCom1)return false;
    if(selComAdic&&p.comisiones.slice(1).indexOf(selComAdic)<0)return false;
    if(selAutor&&p.autores.indexOf(selAutor)<0)return false;
    if(fDesde||fHasta){
      var fp=parseFecha(p.fecha);
      if(fp){
        if(fDesde&&fp<fDesde)return false;
        if(fHasta&&fp>fHasta)return false;
      }
    }
    if(q){
      var hay=(p.extracto+' '+p.autores.join(' ')+' '+p.comisiones.join(' ')).toLowerCase();
      if(hay.indexOf(q)<0)return false;
    }
    return true;
  });
}
function buildCard(p, extraBadge){
  var fg=TIPO_FG[p.tipo]||'#888',bg=TIPO_BG[p.tipo]||'#eee';
  var autoresTxt=p.autores.slice(0,3).join(' \u00b7 ')+(p.autores.length>3?' +'+(p.autores.length-3)+' m\u00e1s':'');
  var btags='',ctags='';
  p.bloques.forEach(function(b){
    var c=getBloqueColor(b);
    btags+='<span class="btag" style="background:'+c+'22;color:'+c+'">'+b+'</span>';
  });
  p.comisiones.forEach(function(c){ctags+='<span class="ctag">'+c+'</span>'});
  var expNro=p.origen+'-'+p.nro+'/'+String(p.anio).slice(-2);
  var linkBtn=p.url?'<a class="exp-link" href="'+p.url+'" target="_blank">Ver expediente &#8599;</a>':'';
  var extra=extraBadge||'';
  return '<div class="card"><div class="card-exp"><div class="exp-id"><span class="exp-badge" style="background:'+bg+';color:'+fg+'">'+p.tipo+'</span><span class="exp-nro">'+expNro+'</span>'+extra+(p.fecha?'<span class="exp-fecha">'+p.fecha+'</span>':'')+'</div>'+linkBtn+'</div><div class="card-body"><div class="extracto">'+p.extracto+'</div><div class="card-meta">'+(autoresTxt?'<div class="meta-row"><span class="meta-bold">'+autoresTxt+'</span></div>':'')+(btags?'<div class="meta-row">'+btags+'</div>':'')+(ctags?'<div class="meta-row">'+ctags+'</div>':'')+'</div></div></div>';
}

function renderList(){
  if(semActivo){
    var tot=semResultados.length;
    document.getElementById('results-count').innerHTML='<span style="color:#1a7a4a;font-weight:600">'+tot+' resultado'+(tot!==1?'s':'')+' sem\u00e1ntico'+(tot!==1?'s':'')+'</span>';
    var html='<div class="sem-banner"><span>Resultados por similitud sem\u00e1ntica</span><button onclick="limpiarSemantico()" style="background:none;border:1px solid #1a7a4a;color:#1a7a4a;border-radius:6px;padding:3px 10px;cursor:pointer;font-size:11px;font-weight:600">Limpiar</button></div>';
    semResultados.forEach(function(r){
      var pct=Math.round(r.score*100);
      var badge='<span class="sem-badge">'+pct+'%</span>';
      html+=buildCard(r.proyecto,badge);
    });
    document.getElementById('list').innerHTML=html;
    return;
  }
  var filtered=getFiltered();
  var tot=filtered.length;
  document.getElementById('results-count').innerHTML=tot+' proyecto'+(tot!==1?'s':'')+' encontrado'+(tot!==1?'s':'');
  if(!filtered.length){
    document.getElementById('list').innerHTML='<div class="no-results">Sin resultados para este filtro.</div>';
    return;
  }
  var html='';
  filtered.forEach(function(p){html+=buildCard(p)});
  document.getElementById('list').innerHTML=html;
}

/* ── Exportar a Excel ──────────────────────────────────────────── */
function exportarExcel(){
  var filtered=getFiltered();
  if(!filtered.length){alert('No hay datos para exportar.');return}
  var headers=['Tipo','Nro','Origen','Fecha','Bloque','Autor','Coautor','Extracto','Giro 1','Giro 2','Giro 3'];
  var rows=[headers];
  var urls=[];
  filtered.forEach(function(p){
    var nroAa=p.nro+'/'+String(p.anio).slice(-2);
    rows.push([
      p.tipo,
      nroAa,
      p.origen,
      p.fecha,
      p.bloques.join('; '),
      p.autores.join('; '),
      (p.coautores||[]).join('; '),
      p.extracto,
      p.comisiones[0]||'',
      p.comisiones[1]||'',
      p.comisiones[2]||''
    ]);
    urls.push(p.url||'');
  });
  var wb=XLSX.utils.book_new();
  var ws=XLSX.utils.aoa_to_sheet(rows);
  /* Hipervínculos en columna Nro */
  for(var i=0;i<filtered.length;i++){
    if(urls[i]){
      var cellRef=XLSX.utils.encode_cell({r:i+1,c:1});
      if(ws[cellRef]){ws[cellRef].l={Target:urls[i]}}
    }
  }
  ws['!cols']=[{wch:6},{wch:10},{wch:8},{wch:12},{wch:28},{wch:35},{wch:35},{wch:60},{wch:30},{wch:30},{wch:30}];
  XLSX.utils.book_append_sheet(wb,ws,'Proyectos');
  XLSX.writeFile(wb,'proyectos_filtrados.xlsx');
}

/* ── Búsqueda semántica ───────────────────────────────────── */
function cosineSim(a,b){var d=0,na=0,nb=0;for(var i=0;i<a.length;i++){d+=a[i]*b[i];na+=a[i]*a[i];nb+=b[i]*b[i]}return d/(Math.sqrt(na)*Math.sqrt(nb)+1e-10)}
async function cargarEmbeddings(){
  if(EMBEDDINGS)return;
  var r=await fetch('embeddings.json');
  EMBEDDINGS=await r.json();
}
async function cargarPipeline(){
  if(semPipeline)return;
  var mod=await import('https://cdn.jsdelivr.net/npm/@xenova/transformers@2.17.2/dist/transformers.min.js');
  mod.env.allowLocalModels=false;
  semPipeline=await mod.pipeline('feature-extraction','Xenova/paraphrase-multilingual-MiniLM-L12-v2');
}
async function buscarSemantico(){
  var q=document.getElementById('sem-search').value.trim();
  if(!q)return;
  var btnEl=document.getElementById('sem-btn');
  var statusEl=document.getElementById('sem-status');
  btnEl.disabled=true;
  try{
    statusEl.textContent='Cargando modelo\u2026 (primera vez ~30s)';
    await cargarEmbeddings();
    await cargarPipeline();
    statusEl.textContent='Calculando similitud\u2026';
    var out=await semPipeline(q,{pooling:'mean',normalize:true});
    var qv=Array.from(out.data);
    var scores=[];
    for(var k in EMBEDDINGS)scores.push({key:k,score:cosineSim(qv,EMBEDDINGS[k])});
    scores.sort(function(a,b){return b.score-a.score;});
    var keyMap={};
    DATA.forEach(function(p){keyMap[p.nro+'-'+p.anio+'-'+p.tipo]=p;});
    semResultados=scores.slice(0,20).map(function(s){return{proyecto:keyMap[s.key],score:s.score};}).filter(function(s){return s.proyecto;});
    semActivo=true;
    statusEl.textContent=semResultados.length+' resultado'+(semResultados.length!==1?'s':'');
    renderList();
  }catch(e){
    statusEl.textContent='Error: '+e.message;
    console.error(e);
  }finally{
    btnEl.disabled=false;
  }
}
function limpiarSemantico(){
  semActivo=false;semResultados=[];
  document.getElementById('sem-search').value='';
  document.getElementById('sem-status').textContent='';
  renderList();
}

/* ── Pestaña Histórico (2010-2024) ──────────────────────────── */
var histActiveAnio='',histActiveTipo='';
function initHist(){
  if(!DATA_HIST||!DATA_HIST.length){
    document.getElementById('hist-count').innerHTML='Sin datos hist\u00f3ricos a\u00fan.';
    return;
  }
  var years={};
  DATA_HIST.forEach(function(p){years[p.anio]=1});
  var ylist=Object.keys(years).map(Number).sort(function(a,b){return b-a});
  var yhtml='<button class="chip on" id="hist-anio-all" onclick="setHistAnio(\'\')">Todos</button>';
  ylist.forEach(function(y){
    yhtml+='<button class="chip" id="hist-anio-'+y+'" onclick="setHistAnio(\''+y+'\')">'+y+'</button>';
  });
  document.getElementById('hist-anio-filters').innerHTML=yhtml;
  var tipos={};
  DATA_HIST.forEach(function(p){tipos[p.tipo]=1});
  var thtml='<button class="chip on" id="hist-tipo-all" onclick="setHistTipo(\'\')">Todos</button>';
  Object.keys(tipos).sort().forEach(function(t){
    thtml+='<button class="chip" id="hist-tipo-'+t+'" onclick="setHistTipo(\''+t+'\')">'+t+' &middot; '+(TIPOS[t]||t)+'</button>';
  });
  document.getElementById('hist-tipo-filters').innerHTML=thtml;
  renderHist();
}
function setHistAnio(anio){
  histActiveAnio=anio;
  document.querySelectorAll('[id^="hist-anio-"]').forEach(function(el){
    var a=el.id.replace('hist-anio-','');
    el.className='chip'+(anio===(a==='all'?'':a)?' on':'');
  });
  renderHist();
}
function setHistTipo(t){
  histActiveTipo=t;
  document.querySelectorAll('[id^="hist-tipo-"]').forEach(function(el){
    var a=el.id.replace('hist-tipo-','');
    el.className='chip'+(t===(a==='all'?'':a)?' on':'');
  });
  renderHist();
}
function getFilteredHist(){
  var q=document.getElementById('hist-search').value.toLowerCase().trim();
  return DATA_HIST.filter(function(p){
    if(histActiveAnio&&String(p.anio)!==histActiveAnio)return false;
    if(histActiveTipo&&p.tipo!==histActiveTipo)return false;
    if(q){
      var hay=(p.extracto+' '+p.autores.join(' ')+' '+p.comisiones.join(' ')).toLowerCase();
      if(hay.indexOf(q)<0)return false;
    }
    return true;
  });
}
function buildCardHist(p){
  var fg=TIPO_FG[p.tipo]||'#888',bg=TIPO_BG[p.tipo]||'#eee';
  var autoresTxt=p.autores.slice(0,3).join(' \u00b7 ')+(p.autores.length>3?' +'+(p.autores.length-3)+' m\u00e1s':'');
  var ctags='';
  p.comisiones.forEach(function(c){ctags+='<span class="ctag">'+c+'</span>'});
  var expNro=p.origen+'-'+p.nro+'/'+String(p.anio).slice(-2);
  var linkBtn=p.url?'<a class="exp-link" href="'+p.url+'" target="_blank">Ver expediente &#8599;</a>':'';
  return '<div class="card"><div class="card-exp"><div class="exp-id"><span class="exp-badge" style="background:'+bg+';color:'+fg+'">'+p.tipo+'</span><span class="exp-nro">'+expNro+'</span>'+(p.fecha?'<span class="exp-fecha">'+p.fecha+'</span>':'')+'</div>'+linkBtn+'</div><div class="card-body"><div class="extracto">'+p.extracto+'</div><div class="card-meta">'+(autoresTxt?'<div class="meta-row"><span class="meta-bold">'+autoresTxt+'</span></div>':'')+(ctags?'<div class="meta-row">'+ctags+'</div>':'')+'</div></div></div>';
}
function renderHist(){
  var filtered=getFilteredHist();
  var tot=filtered.length;
  document.getElementById('hist-count').innerHTML=tot+' proyecto'+(tot!==1?'s':'')+' encontrado'+(tot!==1?'s':'');
  if(!filtered.length){document.getElementById('hist-list').innerHTML='<div class="no-results">Sin resultados para este filtro.</div>';return;}
  var html='';
  filtered.forEach(function(p){html+=buildCardHist(p)});
  document.getElementById('hist-list').innerHTML=html;
}
function exportarHistorial(){
  var filtered=getFilteredHist();
  if(!filtered.length){alert('No hay datos para exportar.');return}
  var rows=[['Tipo','Nro','Origen','Fecha','Autor','Extracto','Giro 1','Giro 2','Giro 3']];
  filtered.forEach(function(p){
    rows.push([p.tipo,p.nro+'/'+String(p.anio).slice(-2),p.origen,p.fecha,
      p.autores.join('; '),p.extracto,p.comisiones[0]||'',p.comisiones[1]||'',p.comisiones[2]||'']);
  });
  var wb=XLSX.utils.book_new();
  var ws=XLSX.utils.aoa_to_sheet(rows);
  ws['!cols']=[{wch:6},{wch:10},{wch:8},{wch:12},{wch:35},{wch:60},{wch:30},{wch:30},{wch:30}];
  XLSX.utils.book_append_sheet(wb,ws,'Historico');
  XLSX.writeFile(wb,'proyectos_historico.xlsx');
}
"""

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Proyectos Ingresados — {titulo}</title>
<link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<style>{css}</style>
</head>
<body>

<div class="header">
  <div class="header-row">
    <span class="header-inst">Senado de la Naci&oacute;n Argentina</span>
    <span class="header-dep">Prosecretar&iacute;a Parlamentaria</span>
  </div>
  <div class="header-title">Proyectos Ingresados</div>
  <div class="header-subtitle">{titulo}</div>
</div>

<div class="tab-bar">
  <button class="tab-btn active" data-tab="dashboard" onclick="switchTab('dashboard')">Dashboard</button>
  <button class="tab-btn" data-tab="detalle" onclick="switchTab('detalle')">Detalle de expedientes</button>
  <button class="tab-btn" data-tab="historico" onclick="switchTab('historico')">Hist&oacute;rico 2010-2024</button>
</div>

<!-- TAB: DASHBOARD -->
<div id="tab-dashboard" class="tab-content active">
  <div class="section-block">
    <div class="section-header">
      <h2>Resumen general</h2>
      <span class="section-hint">Toc&aacute; las barras para filtrar</span>
    </div>
    <div class="section-body">
      <div id="dash-context" class="dash-context"></div>
      <div class="filter-row" style="margin-bottom:12px">
        <button class="chip on" id="dash-anio-all" onclick="setDashAnio('')">Todos</button>
        <button class="chip" id="dash-anio-2025" onclick="setDashAnio('2025')">2025</button>
        <button class="chip" id="dash-anio-2026" onclick="setDashAnio('2026')">2026</button>
      </div>
      <div class="dash-stats-row">
        <div class="stat-card">
          <div class="stat-num" id="stat-total">{total}</div>
          <div class="stat-label">Total proyectos</div>
        </div>
        <div class="stat-card" style="border-left-color:#2E75B6">
          <div class="stat-num" style="color:#2E75B6" id="stat-pl">{pl}</div>
          <div class="stat-label">Proyectos de ley</div>
        </div>
        <div class="stat-card" style="border-left-color:#5B4DA0">
          <div class="stat-num" style="color:#5B4DA0" id="stat-pd">{pd}</div>
          <div class="stat-label">Declaraciones</div>
        </div>
        <div class="stat-card" style="border-left-color:#1a7a4a">
          <div class="stat-num" style="color:#1a7a4a" id="stat-otros">{otros}</div>
          <div class="stat-label">Otros tipos</div>
        </div>
      </div>
      <div class="dash-panels-row">
        <div>
          <div class="dash-subtitle">Por tipo de proyecto</div>
          <div id="tipo-bars"></div>
        </div>
        <div>
          <div class="dash-subtitle">Por bloque pol&iacute;tico</div>
          <div id="bloque-bars"></div>
        </div>
        <div>
          <div class="dash-subtitle">Por comisiones (Top 10)</div>
          <div id="com-bars"></div>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- TAB: DETALLE -->
<div id="tab-detalle" class="tab-content">
  <div class="detalle-layout">
    <div class="filters-panel">
      <div class="section-header">
        <h2>B&uacute;squeda y filtros</h2>
      </div>
      <div class="filters-body">
        <div class="filter-label" style="margin-top:0">A&ntilde;o</div>
        <div class="filter-row" style="margin-bottom:10px">
          <button class="chip on" id="anio-det-all" onclick="setAnioDetalle('')">Todos</button>
          <button class="chip" id="anio-det-2025" onclick="setAnioDetalle('2025')">2025</button>
          <button class="chip" id="anio-det-2026" onclick="setAnioDetalle('2026')">2026</button>
        </div>

        <input class="search-box" type="text" id="search" placeholder="Buscar por extracto, autor o comisi&oacute;n&hellip;" oninput="renderList()">

        <div class="filter-label">Tipo</div>
        <div class="filter-row" id="tipo-filters"></div>

        <div class="filter-label">Bloque</div>
        <div class="select-wrapper">
          <select class="filter-select" id="bloque-select" onchange="setBloque(this.value)">
            <option value="">Todos los bloques</option>
          </select>
          <span class="select-arrow">&#9660;</span>
        </div>

        <div class="filter-label">Provincia</div>
        <div class="select-wrapper">
          <select class="filter-select" id="provincia-select" onchange="setProvincia(this.value)">
            <option value="">Todas las provincias</option>
          </select>
          <span class="select-arrow">&#9660;</span>
        </div>

        <div class="filter-label">Origen</div>
        <div class="filter-row" id="origen-filters"></div>

        <div class="filter-label">Rango de fechas</div>
        <div class="date-range">
          <input type="date" class="date-input" id="fecha-desde" onchange="renderList()">
          <span class="date-sep">hasta</span>
          <input type="date" class="date-input" id="fecha-hasta" onchange="renderList()">
        </div>

        <div class="filter-label">Comisi&oacute;n (1er giro)</div>
        <div class="select-wrapper">
          <select class="filter-select" id="com-select-1" onchange="renderList()">
            <option value="">Todas las comisiones</option>
          </select>
          <span class="select-arrow">&#9660;</span>
        </div>

        <div class="filter-label">Comisi&oacute;n (giros adicionales)</div>
        <div class="select-wrapper">
          <select class="filter-select" id="com-select-adic" onchange="renderList()">
            <option value="">Todos los giros adicionales</option>
          </select>
          <span class="select-arrow">&#9660;</span>
        </div>

        <div class="filter-label">Autor</div>
        <div class="select-wrapper">
          <select class="filter-select" id="autor-select" onchange="renderList()">
            <option value="">Todos los autores</option>
          </select>
          <span class="select-arrow">&#9660;</span>
        </div>

        <!-- SEMÁNTICO: desactivado temporalmente - reactivar cuando se retome
        <hr class="sem-divider" id="sem-wrap" style="display:none;margin-top:14px">
        <div id="sem-inner" style="display:none">
          <div style="display:flex;align-items:center;gap:6px;margin-bottom:4px;margin-top:12px">
            <span class="filter-label" style="margin:0">B&uacute;squeda sem&aacute;ntica</span>
            <span style="font-size:9px;font-weight:700;background:#EDE8FA;color:#5B4DA0;padding:2px 7px;border-radius:10px;letter-spacing:.5px">BETA</span>
          </div>
          <div style="font-size:10px;color:#aaa;font-style:italic;margin-bottom:7px;line-height:1.4">Buscá por concepto, no por palabra exacta.</div>
          <div style="display:flex;gap:6px;margin-bottom:4px">
            <input class="search-box" type="text" id="sem-search" placeholder="Ej: proyectos sobre educaci&oacute;n ambiental&hellip;" style="margin-bottom:0" onkeydown="if(event.key==='Enter')buscarSemantico()">
            <button class="btn-sem" id="sem-btn" onclick="buscarSemantico()">Buscar</button>
          </div>
          <div class="sem-status" id="sem-status"></div>
        </div>
        -->
      </div>
    </div>

    <div class="results-panel">
      <div class="results-header">
        <span class="results-count" id="results-count"></span>
        <button class="btn-export" onclick="exportarExcel()">&#128196; Exportar Excel</button>
      </div>
      <div id="list"></div>
    </div>
  </div>
</div>

<!-- TAB: HISTÓRICO -->
<div id="tab-historico" class="tab-content">
  <div class="section-block">
    <div class="section-header">
      <h2>Expedientes hist&oacute;ricos</h2>
      <span class="section-hint">2010–2024 &middot; sin info de bloque pol&iacute;tico</span>
    </div>
    <div class="section-body">
      <div id="hist-anio-filters" class="filter-row" style="margin-bottom:10px"></div>
      <div id="hist-tipo-filters" class="filter-row" style="margin-bottom:10px"></div>
      <input class="search-box" type="text" id="hist-search" placeholder="Buscar por extracto, autor o comisi&oacute;n&hellip;" oninput="if(histInited)renderHist()">
      <div class="results-header">
        <span class="results-count" id="hist-count"></span>
        <button class="btn-export" onclick="exportarHistorial()">&#128196; Exportar Excel</button>
      </div>
      <div id="hist-list"></div>
    </div>
  </div>
</div>

<div class="footer">Prosecretar&iacute;a Parlamentaria &middot; Senado de la Naci&oacute;n Argentina<br>Datos al {fecha}</div>

<script>
var DATA = {datos};
{js}
init();
</script>
</body>
</html>"""


def parse_fecha_sort(fecha_str):
    # Convierte 'DD/MM/AAAA' a 'AAAAMMDD' para que Python pueda ordenar cronológicamente
    if not fecha_str: return "00000000"
    parts = fecha_str.split("/")
    if len(parts) == 3:
        return f"{parts[2]}{parts[1]}{parts[0]}"
    return "00000000"

def generar_desde_lista(proyectos, titulo_periodo, fecha_datos, archivo_salida="index.html", embeddings_path="embeddings.json", proyectos_hist=None):
    # Ordena por fecha exacta, luego por año y finalmente por número de expediente (más nuevo a más viejo)
    proyectos = sorted(proyectos, key=lambda x: (parse_fecha_sort(x.get("fecha", "")), x["anio"], x["nro"]), reverse=True)
    total = len(proyectos)
    tipos_count = {}
    for p in proyectos:
        tipos_count[p["tipo"]] = tipos_count.get(p["tipo"], 0) + 1
    datos_js = json.dumps(proyectos, ensure_ascii=False)
    hist_js = json.dumps(sorted(proyectos_hist or [], key=lambda x: (x["anio"], x["nro"]), reverse=True), ensure_ascii=False)
    # SEMÁNTICO: desactivado temporalmente - reactivar cuando se retome
    js_final = f"var HAS_EMBEDDINGS=false;\nvar DATA_HIST={hist_js};\n" + JS
    html_final = HTML_TEMPLATE.format(
        titulo = titulo_periodo,
        fecha  = fecha_datos,
        total  = total,
        pl     = tipos_count.get("PL", 0),
        pd     = tipos_count.get("PD", 0),
        otros  = total - tipos_count.get("PL", 0) - tipos_count.get("PD", 0),
        css    = CSS,
        datos  = datos_js,
        js     = js_final,
    )
    with open(archivo_salida, "w", encoding="utf-8") as f:
        f.write(html_final)
    print(f"Listo. Archivo generado: {archivo_salida}  ({len(html_final):,} bytes)")
    print(f"  → {total} proyectos  |  {tipos_count.get('PL',0)} PL  |  {tipos_count.get('PD',0)} PD")


if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("ERROR: falta openpyxl. pip install openpyxl")
        sys.exit(1)

    print(f"Leyendo senadores desde: {EXCEL_SENADORES}")
    try:
        wb_sen = openpyxl.load_workbook(EXCEL_SENADORES)
    except FileNotFoundError:
        print(f"ERROR: no se encontró '{EXCEL_SENADORES}'")
        sys.exit(1)
    ws_sen = wb_sen.active
    senador_bloque = {}
    for row in ws_sen.iter_rows(min_row=2, values_only=True):
        bloque, apellido, nombre = row[0], row[1], row[2]
        if apellido and nombre:
            key = f"{apellido.strip()}, {nombre.strip()}".upper()
            senador_bloque[key] = bloque
    print(f"  → {len(senador_bloque)} senadores cargados")

    print(f"Leyendo proyectos desde: {EXCEL_PROYECTOS}")
    try:
        wb_proy = openpyxl.load_workbook(EXCEL_PROYECTOS)
    except FileNotFoundError:
        print(f"ERROR: no se encontró '{EXCEL_PROYECTOS}'")
        sys.exit(1)
    ws_proy = wb_proy.active
    headers = [cell.value for cell in ws_proy[1]]
    nro_links = {}
    for row in ws_proy.iter_rows(min_row=2):
        cell = row[1]
        if cell.value and cell.hyperlink:
            url = cell.hyperlink.target if hasattr(cell.hyperlink, "target") else str(cell.hyperlink)
            nro_links[int(cell.value)] = url

    def parse_autores(s):
        if not s: return []
        return [p.strip().rstrip("-").strip() for p in s.split(" - ") if p.strip().rstrip("-").strip()]

    def get_bloques_excel(autores):
        seen, result = set(), []
        for a in autores:
            b = senador_bloque.get(a.upper(), "Sin datos")
            if b not in seen: seen.add(b); result.append(b)
        return result

    proyectos = []
    for row in ws_proy.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        if not any(v for v in row): continue
        autores    = parse_autores(r.get("AUTOR", ""))
        bloques    = get_bloques_excel(autores)
        comisiones = [r.get(f"COMISION{i}") for i in range(1, 4) if r.get(f"COMISION{i}")]
        mesa       = r.get("MESA DE ENTRADAS", "") or ""
        fecha      = mesa.split(" -")[0].strip() if mesa else ""
        caratula   = r.get("CARÁTULA", "") or ""
        extracto   = caratula[caratula.index(":") + 1:].strip() if ":" in caratula else caratula.strip()
        nro        = int(r["NRO."]) if r.get("NRO.") else 0
        origen     = r.get("ORIGEN", "S") or "S"
        proyectos.append({
            "nro": nro, "anio": int(r["AÑO"]) if r.get("AÑO") else 2026,
            "tipo": r.get("TIPO", ""), "tipo_label": TIPOS.get(r.get("TIPO", ""), r.get("TIPO", "")),
            "extracto": extracto, "autores": autores, "coautores": [], "bloques": bloques,
            "provincias": [], "comisiones": comisiones, "fecha": fecha,
            "dae": r.get("NRO. DAE / DADO CUENTA", "") or "",
            "origen": origen, "url": nro_links.get(nro, ""),
        })

    print(f"  → {len(proyectos)} proyectos procesados")
    generar_desde_lista(proyectos, TITULO_PERIODO, FECHA_DATOS, ARCHIVO_SALIDA)
